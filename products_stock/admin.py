from products_stock.models import Publisher, Products
from django.contrib import admin
from django.urls import path
from django.shortcuts import render
from products_stock.product_class import ProductClass
from zipfile import ZipFile
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, NamedStyle,\
    PatternFill
import os
from bg_store.settings import BASE_DIR
from django.views.static import serve
from products_stock.forms.financial_data_form import FinancialForm


@admin.register(Publisher)
class PublisherAdmin(admin.ModelAdmin):
    ordering = ('name',)


def import_excel(request):
    if "GET" == request.method:

        return render(request, 'admin/import_excel.html')
    else:
        excel_file = request.FILES['excel_file']
        zip_file = request.FILES['ZIP-images']
        bg_cover_root = os.path.join(BASE_DIR, 'media/bg_cover')
        wb = load_workbook(excel_file,
                           data_only=True)

        sheet = wb.active

        products_list = [row for row in
                         sheet.iter_rows(min_row=2, max_col=14,
                                         values_only=True) if
                         None not in row]
        for product in products_list:
            new_prod = ProductClass(product)
            new_prod.verify_prod()

        with ZipFile(zip_file, 'r') as zipObj:
            zipObj.extractall(bg_cover_root)

    return render(request, 'admin/import_excel.html')


def download_template(request):

    filepath = 'util/template_form_data.xlsx'

    return serve(request, os.path.basename(filepath), os.path.dirname(filepath))




@admin.register(Products)
class ProductsAdmin(admin.ModelAdmin):
    list_display = ('name',
                    'publisher',
                    'price',
                    'new_stock',
                    'used_stock',
                    'sales',
                    'created_at',
                    'updated_at')
    ordering = ('name',)
    # date_hierarchy = 'updated_at'

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [

            path('import/', self.admin_site.admin_view(import_excel),
                 name='import'
                 ),
            path('download_template/', self.admin_site.admin_view(
                download_template),
                 name='download_template',
                 ),
            path('financial_data/', self.admin_site.admin_view(
                financial_view),
                 name='financial-data',
                 ),
        ]
        return my_urls + urls


def financial_view(request):
    # Once submitted, if executes.
    if request.method == 'POST':
        form = FinancialForm(request.POST)
        if form.is_valid():

            order_item_dict = form.retrieve_order_items_date_range()

            # excel part
            excel_file = os.path.join(BASE_DIR, 'util/financial_data.xlsx')

            workbook = Workbook()  # creating workbook object
            sheet = workbook.active  # selecting active sheet( sheet 1)

            sheet.append(['item_id', 'product_name', 'quantity',
                          'SKU', 'price', 'stock', 'sales'
                          ])
            # styles
            header = NamedStyle(name='header')
            double_border_side = Side(border_style="double")
            square_border = Border(
                top=double_border_side,
                right=double_border_side,
                bottom=double_border_side,
                left=double_border_side,
            )
            header.font = Font(bold=True)
            header.border = square_border
            header.alignment = Alignment(horizontal="center")
            header_row = sheet[1]
            for cell in header_row:
                cell.style = header
                cell.fill = PatternFill(start_color="0000FF00",
                                        end_color="0000FF00",
                                        fill_type="solid"
                                        )
            # populating the excel sheet
            for prod_id, quantity in order_item_dict.items():
                product = Products.objects.get(id=prod_id)
                data = [
                    prod_id,  # column A
                    product.name,  # column B
                    quantity,  # column C
                    product.sku,  # column D
                    product.price,  # column E
                    product.new_stock,  # column F
                    product.sales  # column G
                ]
                sheet.append(data)
            # each row will be a product

            workbook.save(filename=excel_file)
            filepath = 'util/financial_data.xlsx'
            return serve(request, os.path.basename(filepath),
                         os.path.dirname(filepath))

    else:
        # GET
        form = FinancialForm()

    return render(request, 'admin/financial_page.html', {
        'form': form,
    })


